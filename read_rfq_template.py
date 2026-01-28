#!/usr/bin/env python3
import sys
import os
import json

# Add temp_packages to path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'temp_packages'))

import openpyxl

excel_file = '/Users/ouyangheng/Desktop/trade/chainsawchain/factory RFQ/RFQ Template.xlsx'
output_json = '/Users/ouyangheng/coding/chainsaw-chain-supplier-nextjs/rfq_template_data.json'

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

# Find header row
header_row_idx = None
for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
    if row and isinstance(row[0], (int, str)) and str(row[0]).strip():
        # Check if this looks like a header row
        if any('No.' in str(cell) if cell else False for cell in row[:3]):
            header_row_idx = idx
            break

if not header_row_idx:
    # Try first row
    header_row_idx = 1

# Get headers - use values_only to get actual values
headers = [str(cell).strip() if cell else '' for cell in list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]]

# Get all data rows
data_rows = []
for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row, values_only=True):
    if row and row[0] is not None and str(row[0]).strip() and str(row[0]) != 'None':
        data_row = [str(cell).strip() if cell and str(cell) != 'None' else '' for cell in row]
        data_rows.append(data_row)

# Parse data into structured format
products = []
for row in data_rows:
    if len(row) >= len(headers):
        product = {}
        for i, header in enumerate(headers):
            if i < len(row):
                product[header] = row[i]
        products.append(product)

# Save as JSON
with open(output_json, 'w', encoding='utf-8') as f:
    json.dump({
        'headers': headers,
        'products': products
    }, f, ensure_ascii=False, indent=2)

print(f"JSON file generated: {output_json}")
print(f"Headers: {headers}")
print(f"Products: {len(products)}")
print("\nFirst product:")
if products:
    print(json.dumps(products[0], ensure_ascii=False, indent=2))
